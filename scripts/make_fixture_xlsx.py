#!/usr/bin/env python3
"""Generate a synthetic 3-sheet fixture workbook (FINDINGS / DEPENDENCIES /
SUMMARY) with random-but-seeded values, print the ground-truth JSON computed
from the generated rows, and emit repo_package_remaining.csv for the
leverage model.

Usage: python3 scripts/make_fixture_xlsx.py OUT.xlsx [--seed 7]
"""

from __future__ import annotations

import argparse
import csv
import json
import random
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

CATEGORIES = ["W-089", "W-079", "W-639", "W-798", "W-862", "W-022",
              "W-352", "W-611", "W-532", "W-918"]
STATUSES = ["Open", "Resolved", "False Positive", "Investigating"]
COMPONENTS = ["libjson", "libyaml", "libpool", "libnet", "liblog",
              "libhttp", "libauth", "libxml", "libcache", "libqueue",
              "libcsv", "libimg", "libpdf", "libsql", "liblegacy"]


def semver(major, minor, patch):
    return f"{major}.{minor}.{patch}"


def build(seed: int):
    rnd = random.Random(seed)
    entities = [f"APP-{i:03d}" for i in range(1, 41)]

    findings = []
    fid = 0
    for rank, ent in enumerate(entities, start=1):
        n = max(3, int(400 / rank ** 0.9 * rnd.uniform(0.7, 1.3)))
        for _ in range(n):
            fid += 1
            cat = rnd.choices(CATEGORIES, weights=range(len(CATEGORIES), 0, -1))[0]
            status = rnd.choices(STATUSES, weights=[50, 42, 5, 3])[0]
            reopened = f"2026-0{rnd.randint(1, 6)}-15" if rnd.random() < 0.06 else ""
            findings.append([f"F-{fid:06d}", ent, "", cat, status,
                             rnd.choices([1, 2, 3], weights=[5, 80, 15])[0],
                             reopened, f"2026-0{rnd.randint(1, 6)}-{rnd.randint(1, 28):02d}"])
    # planted edge rows: secondary-only id, and 2 rows with no entity at all
    fid += 1
    findings.append([f"F-{fid:06d}", "", "MF-117", "W-079", "Open", 2, "", "2026-06-01"])
    for _ in range(2):
        fid += 1
        findings.append([f"F-{fid:06d}", "", "", "W-022", "Open", 2, "", "2026-07-11"])
    # planted: APP-030 heavily resolved (resolved > open)
    for _ in range(60):
        fid += 1
        findings.append([f"F-{fid:06d}", "APP-030", "", rnd.choice(CATEGORIES),
                         "Resolved", 2, "", "2025-11-05"])

    repos = [f"git/team-{c}/svc-{s}" for c in "abcde"
             for s in ("billing", "auth", "etl", "web", "api")]
    deps = []
    for r_i, repo in enumerate(repos, start=1):
        ent = rnd.choice(entities)
        for c_i, comp in enumerate(COMPONENTS, start=1):
            if rnd.random() > 0.5:
                continue
            cur = semver(rnd.randint(1, 6), rnd.randint(0, 20), rnd.randint(0, 9))
            parts = [int(x) for x in cur.split(".")]
            tgt = semver(parts[0] + (1 if rnd.random() < 0.25 else 0),
                         parts[1] + rnd.randint(0, 3), rnd.randint(0, 9))
            comp_cell = f"com.ex/{comp}@{cur}" if rnd.random() < 0.5 else comp
            resolved = "Yes" if rnd.random() < 0.45 else "No"
            weight = max(1, int(120 / (r_i ** 0.7) / (c_i ** 0.4) * rnd.uniform(0.5, 1.5)))
            deps.append([repo, ent, comp_cell, cur, tgt, resolved, weight])
    # planted edge rows (all unresolved so they surface)
    deps.append(["git/team-a/svc-billing", "APP-001", "libyaml", "2.5", "2.5", "No", 7])          # harmonize
    deps.append(["git/team-d/svc-legacy", "APP-011", "liblegacy", "2.6", "1.0.1", "No", 2])       # review: target<current
    deps.append(["git/team-c/svc-etl", "APP-009", "libnet", "4.2.0.RC4", "4.2.16.Final", "No", 5])  # review: pre-release
    deps.append(["git/team-b/svc-auth", "APP-004", "libpool", "1.0.0.Alpha3", "1.0.9", "No", 4])  # review: pre-release
    deps.append(["git/team-e/svc-api", "APP-020", "libhttp", "3.0.0-M2", "3.1.0", "No", 3])       # review: pre-release
    return findings, deps


def truths(findings, deps):
    open_rows = [f for f in findings if f[4] == "Open"]
    by_ent = defaultdict(int)
    for f in open_rows:
        ent = f[1] or f[2]
        if ent:
            by_ent[ent] += 1
    top_ent = max(by_ent.items(), key=lambda kv: kv[1])
    unresolved = [d for d in deps if d[5] == "No"]
    remaining_by_repo = defaultdict(int)
    for d in unresolved:
        remaining_by_repo[d[0]] += 1
    top_repo = max(remaining_by_repo.items(), key=lambda kv: kv[1])
    occ_total = len(findings) + sum(d[6] for d in deps)
    return {
        "findings_rows": len(findings),
        "open_findings": len(open_rows),
        "open_rows_without_entity": sum(1 for f in open_rows if not (f[1] or f[2])),
        "top_entity": {"id": top_ent[0], "open": top_ent[1]},
        "dependencies_rows": len(deps),
        "unresolved_dep_rows": len(unresolved),
        "occurrence_total": occ_total,
        "top_repo_by_unresolved_rows": {"repo": top_repo[0], "rows": top_repo[1]},
        "planted": {
            "harmonize": "libyaml 2.5 -> 2.5",
            "review_downgrade": "liblegacy 2.6 -> 1.0.1",
            "review_prerelease": ["libnet 4.2.0.RC4", "libpool 1.0.0.Alpha3",
                                  "libhttp 3.0.0-M2"],
            "resolved_exceeds_open_entity": "APP-030",
        },
    }


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("out", type=Path)
    ap.add_argument("--seed", type=int, default=7)
    args = ap.parse_args()

    findings, deps = build(args.seed)
    truth = truths(findings, deps)

    wb = Workbook()
    ws = wb.active
    ws.title = "FINDINGS"
    ws.append(["finding_id", "entity_primary", "entity_secondary", "category",
               "status", "sla_class", "reopened_date", "found_date"])
    for row in findings:
        ws.append(row)

    ws2 = wb.create_sheet("DEPENDENCIES")
    ws2.append(["repo", "entity_id", "component", "current_version",
                "target_version", "resolved", "row_count"])
    for row in deps:
        ws2.append(row)

    ws3 = wb.create_sheet("SUMMARY")
    for key in ("findings_rows", "open_findings", "dependencies_rows",
                "unresolved_dep_rows", "occurrence_total"):
        ws3.append([key, truth[key]])

    wb.save(args.out)

    # leverage-model input: unresolved rows grouped by repo x base component
    grouped = defaultdict(int)
    for d in deps:
        if d[5] == "No":
            base = d[2].split("@")[0]
            grouped[(d[0], base)] += 1
    csv_path = args.out.with_name("repo_package_remaining.csv")
    with csv_path.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["repo", "package", "remaining"])
        for (repo, pkg), n in sorted(grouped.items(), key=lambda kv: -kv[1]):
            w.writerow([repo, pkg, n])

    print(f"wrote {args.out} and {csv_path}")
    print(json.dumps(truth, indent=2))


if __name__ == "__main__":
    main()
